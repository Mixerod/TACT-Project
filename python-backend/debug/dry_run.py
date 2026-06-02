"""
dry_run.py -- simulate a full process run WITHOUT writing any file.

Use to confirm the mapping is correct before running for real: it reports the
exact value each mapping would write and the exact cell it would land in
(including merged-cell redirects).

Usage:
    python debug/dry_run.py "C:/path/to/file.csv" "C:/path/to/profile.json"
"""
import sys
from pathlib import Path

from _common import (
    safe_print, section, die, require_file, load_profile_dict, build_dataframe,
)


def _effective_target(ws, address):
    """Mirror write_to_cell_safe's redirect: merged child -> range top-left."""
    from openpyxl.utils import coordinate_to_tuple, get_column_letter

    if type(ws[address]).__name__ != "MergedCell":
        return address
    row, col = coordinate_to_tuple(address)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return f"{get_column_letter(rng.min_col)}{rng.min_row}"
    return address


def main(argv):
    if len(argv) < 3:
        die("usage: python debug/dry_run.py <csv> <profile.json>")

    csv_path = require_file(argv[1], "data file")
    profile = load_profile_dict(argv[2])

    from openpyxl import load_workbook
    from services import excel_writer
    from services.matcher import resolve_identity, find_existing_report, build_output_filename

    template_path = require_file(profile["template"]["path"], "template")
    sheet_name = profile["template"]["sheet_name"]
    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        die(f"sheet '{sheet_name}' not in template (have: {wb.sheetnames})")
    ws = wb[sheet_name]

    # --- identity + dataframe (same as the real pipeline) ---
    identity = resolve_identity(csv_path, profile.get("identity", {}))
    section("IDENTITY")
    safe_print(f"  order={identity['order']!r} color={identity['color']!r} "
               f"confidence={identity['confidence']} warnings={identity['warnings']}")
    if not identity.get("order"):
        safe_print("  NOTE: empty order -> the real run would FAIL LOUD before writing.")

    header_row = profile.get("source", {}).get("header_row", 0)
    df = build_dataframe(csv_path, header_row)
    safe_print(f"  data rows={len(df)} columns={list(df.columns)}")

    # --- output file resolution (no file is created) ---
    section("OUTPUT FILE (not created)")
    out_dir = profile["output"]["directory"]
    existing = find_existing_report(out_dir, identity["order"], identity["color"])
    if existing:
        safe_print(f"  would REUSE existing: {existing}")
    else:
        name = build_output_filename(
            profile["output"]["filename_pattern"], identity["order"], identity["color"],
            profile["output"].get("date_format", "YYYYMMDD"), profile.get("method_code", ""),
        )
        safe_print(f"  would CREATE new   : {Path(out_dir) / name}")

    # --- record writes instead of performing them ---
    records = []  # (address, effective, value)

    def recorder(ws_, address, value):
        records.append((address, _effective_target(ws_, address), value))

    original = excel_writer.write_to_cell_safe
    excel_writer.write_to_cell_safe = recorder
    try:
        section("WRITES (simulated)")

        # 1. identity cells
        _run("identity", records, lambda: excel_writer.apply_identity_mapping(
            ws, identity, profile.get("identity", {}).get("output_cells", [])))

        # 2-4. each configured mapping
        for m in profile.get("mappings", []):
            label = f"{m.get('type')} [{m.get('label', m.get('id'))}]"
            _run(label, records, lambda m=m: _apply(excel_writer, ws, df, m, csv_path))
    finally:
        excel_writer.write_to_cell_safe = original

    wb.close()


def _apply(excel_writer, ws, df, mapping, csv_path):
    t = mapping.get("type")
    if t == "column":
        excel_writer.apply_column_mapping(ws, df, mapping)
    elif t == "cell":
        excel_writer.apply_cell_mapping(ws, df, mapping, csv_path=csv_path)
    elif t == "range":
        excel_writer.apply_range_mapping(ws, df, mapping)
    else:
        raise ValueError(f"unknown mapping type: {t!r}")


def _run(label, records, fn):
    """Run one mapping, print the writes it produced, and report errors inline."""
    before = len(records)
    try:
        fn()
    except Exception as e:  # noqa: BLE001 - debug tool: report, do not abort the run
        safe_print(f"  [{label}] ERROR: {getattr(e, 'message', None) or e}")
        return
    produced = records[before:]
    if not produced:
        safe_print(f"  [{label}] (no cells written)")
        return
    for address, effective, value in produced:
        redirect = "" if address == effective else f"  (merged -> {effective})"
        safe_print(f"  [{label}] {address} = {value!r}{redirect}")


if __name__ == "__main__":
    main(sys.argv)
