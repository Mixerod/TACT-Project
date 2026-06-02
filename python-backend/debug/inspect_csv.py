"""
inspect_csv.py -- dump the real structure of a TACT data file.

Use when a labtech reports "the file won't read": run this to see what the file
actually contains.

Usage:
    python debug/inspect_csv.py "C:/path/to/file.csv"
    python debug/inspect_csv.py "C:/path/to/file.xlsx"

Prints:
    - all sheet names
    - headers of the 'Result' sheet (or the data table for a raw CSV)
    - the first 5 data rows
    - the value found next to 'Sample name' / 'Submission' in the Condition block
"""
import sys
from pathlib import Path

from _common import safe_print, section, die, require_file, read_csv_rows


CONDITION_KEYS = ["Sample name", "Submission"]


def _find_next_to(rows, key):
    """Return the cell immediately to the right of `key`, scanning a 2D grid."""
    key_l = key.strip().lower()
    for row in rows:
        for c in range(len(row) - 1):
            if str(row[c]).strip().lower() == key_l:
                val = str(row[c + 1]).strip()
                if val:
                    return val
    return None


def inspect_excel(path):
    import pandas as pd
    from openpyxl import load_workbook

    xl = pd.ExcelFile(path)
    section("SHEET NAMES")
    for s in xl.sheet_names:
        safe_print(" -", s)

    result_sheet = "Result" if "Result" in xl.sheet_names else xl.sheet_names[0]
    df = pd.read_excel(path, sheet_name=result_sheet)
    section(f"HEADERS of sheet '{result_sheet}'")
    safe_print(list(df.columns))
    section("FIRST 5 ROWS")
    safe_print(df.head(5).to_string(index=False))

    section("CONDITION lookup")
    if "Condition" in xl.sheet_names:
        wb = load_workbook(path, data_only=True)
        ws = wb["Condition"]
        grid = [[c.value for c in row] for row in ws.iter_rows()]
        wb.close()
        for key in CONDITION_KEYS:
            safe_print(f"  {key!r} -> {_find_next_to(grid, key)!r}")
    else:
        safe_print("  (no 'Condition' sheet in this workbook)")


def inspect_csv(path):
    rows, encoding = read_csv_rows(path)
    section("FILE INFO")
    safe_print(f"  encoding : {encoding}")
    safe_print(f"  raw rows : {len(rows)}")
    safe_print("  logical sheets: ['Result']  (a .csv is a single Result sheet)")

    # Heuristic: the data-table header is the widest of the first several rows.
    scan = rows[:8]
    header_idx = max(range(len(scan)), key=lambda i: len(scan[i])) if scan else 0
    section(f"DATA TABLE HEADERS (row index {header_idx})")
    safe_print([str(c).strip() for c in rows[header_idx]])

    section("FIRST 5 DATA ROWS")
    for r in rows[header_idx + 1: header_idx + 6]:
        safe_print(" ", r)

    section("CONDITION lookup (scanning all rows)")
    for key in CONDITION_KEYS:
        safe_print(f"  {key!r} -> {_find_next_to(rows, key)!r}")


def main(argv):
    if len(argv) < 2:
        die("usage: python debug/inspect_csv.py <path-to-csv-or-xlsx>")
    path = require_file(argv[1], "data file")
    safe_print(f"Inspecting: {path}")

    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        inspect_excel(path)
    elif ext == ".csv":
        inspect_csv(path)
    else:
        die(f"unsupported extension '{ext}' (expected .csv/.xlsx/.xls)")


if __name__ == "__main__":
    main(sys.argv)
