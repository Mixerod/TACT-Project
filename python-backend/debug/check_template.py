"""
check_template.py -- validate a profile's target cells against a real template.

Use to catch mapping mistakes before processing: are all destination cells real,
do any land on merged cells, and do two mappings fight over the same cell?

Usage:
    python debug/check_template.py "C:/path/to/template.xlsx" "C:/path/to/profile.json"

Prints, per destination:
    - the declared cell, whether it parses, and whether it's inside the sheet bounds
    - the effective cell after merged-range redirect
And a final CONFLICTS section listing any effective cell written by 2+ mappings.
"""
import sys

from _common import safe_print, section, die, require_file, load_profile_dict


def _effective(ws, address):
    from openpyxl.utils import coordinate_to_tuple, get_column_letter

    if type(ws[address]).__name__ != "MergedCell":
        return address, False
    row, col = coordinate_to_tuple(address)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return f"{get_column_letter(rng.min_col)}{rng.min_row}", True
    return address, True


def _destinations(profile):
    """Yield (label, declared_cell) for every cell a profile would write to."""
    ident = profile.get("identity", {})
    for oc in ident.get("output_cells", []):
        if oc.get("cell"):
            yield f"identity:{oc.get('field')}", oc["cell"]

    for m in profile.get("mappings", []):
        label = f"{m.get('type')}:{m.get('label', m.get('id'))}"
        t = m.get("type")
        if t == "cell" and m.get("excel_cell"):
            yield label, m["excel_cell"]
        elif t == "column" and m.get("excel_column") and m.get("excel_start_row"):
            # The column write starts here and extends downward per data row.
            yield f"{label} (column start)", f"{m['excel_column']}{m['excel_start_row']}"
        elif t == "range" and m.get("excel_start_cell"):
            yield f"{label} (range start)", m["excel_start_cell"]


def main(argv):
    if len(argv) < 3:
        die("usage: python debug/check_template.py <template.xlsx> <profile.json>")

    template_path = require_file(argv[1], "template")
    profile = load_profile_dict(argv[2])

    from openpyxl import load_workbook
    from openpyxl.utils import coordinate_to_tuple

    wb = load_workbook(template_path)
    sheet_name = profile.get("template", {}).get("sheet_name")
    if sheet_name not in wb.sheetnames:
        die(f"sheet '{sheet_name}' not in template (have: {wb.sheetnames})")
    ws = wb[sheet_name]
    max_row, max_col = ws.max_row, ws.max_column

    safe_print(f"Template : {template_path}")
    safe_print(f"Sheet    : {sheet_name}  (bounds {max_row} rows x {max_col} cols)")

    section("DESTINATIONS")
    seen = {}      # effective cell -> list of labels
    problems = 0
    for label, declared in _destinations(profile):
        try:
            row, col = coordinate_to_tuple(declared)
        except Exception:
            safe_print(f"  [{label}] {declared}: INVALID cell address")
            problems += 1
            continue

        in_bounds = row <= max_row and col <= max_col
        eff, merged = _effective(ws, declared)
        flags = []
        if not in_bounds:
            flags.append("OUT-OF-BOUNDS")
            problems += 1
        if merged:
            flags.append(f"merged -> {eff}")
        suffix = ("  [" + ", ".join(flags) + "]") if flags else "  [ok]"
        safe_print(f"  [{label}] {declared}{suffix}")
        seen.setdefault(eff, []).append(label)

    section("CONFLICTS (same effective cell written by 2+ mappings)")
    conflicts = {cell: labels for cell, labels in seen.items() if len(labels) > 1}
    if not conflicts:
        safe_print("  none")
    else:
        for cell, labels in conflicts.items():
            safe_print(f"  {cell} <- {labels}")

    section("SUMMARY")
    safe_print(f"  destinations checked : {sum(len(v) for v in seen.values())}")
    safe_print(f"  problems             : {problems}")
    safe_print(f"  conflicts            : {len(conflicts)}")
    wb.close()
    sys.exit(1 if (problems or conflicts) else 0)


if __name__ == "__main__":
    main(sys.argv)
