"""
Tests for services.excel_writer against the real TACT Excel template.

The real template is heavily merged, so writes to a merged child cell are
redirected to the range's top-left cell. Verifications therefore read the
*effective* value via the ``effective_reader`` helper (see conftest).
"""
import os
from datetime import datetime

import pandas as pd
import pytest
from openpyxl import load_workbook

from services.excel_writer import (
    prepare_output_file,
    write_excel_safe,
    apply_column_mapping,
    apply_cell_mapping,
    apply_identity_mapping,
)


class MockIdentity:
    def __init__(self, order, color):
        self.order = order
        self.color = color


@pytest.fixture
def profile_factory(template_path, template_sheet):
    """Build a dict profile pointed at the real template and a given output dir."""
    def _make(output_dir):
        return {
            "id": "excel-writer-test",
            "name": "Excel Writer Test Profile",
            "method_code": "tensile",
            "template": {"path": template_path, "sheet_name": template_sheet},
            "output": {
                "directory": str(output_dir),
                "filename_pattern": "Report_{order}_{color}_{date}.xlsx",
                "date_format": "YYYYMMDD",
            },
            "mappings": [
                {  # 0: column -> C15 redirects to top-left C10
                    "id": "col-force", "type": "column", "label": "Max Force",
                    "csv_column": "Max Force (N)", "excel_column": "C", "excel_start_row": 15,
                },
                {  # 1: system date -> B7 redirects to top-left A7
                    "id": "cell-date", "type": "cell", "label": "Date",
                    "value_source": "system_date", "excel_cell": "B7",
                },
                {  # 2: static -> B12 redirects to top-left B10
                    "id": "cell-static", "type": "cell", "label": "Spec",
                    "value_source": "static:ISO 13934-1", "excel_cell": "B12",
                },
            ],
            "identity": {
                "output_cells": [
                    {"field": "order", "cell": "B8"},   # -> top-left A8
                    {"field": "color", "cell": "B9"},   # plain cell
                ]
            },
        }
    return _make


def test_column_mapping(tmp_path, profile_factory, effective_reader):
    """A column value lands at the effective (top-left) target cell."""
    profile = profile_factory(tmp_path)
    df = pd.DataFrame({"Sample ID": ["S-1"], "Max Force (N)": [245.3]})
    identity = MockIdentity("ORD-C-1", "RED")

    output_path = prepare_output_file(profile, identity)
    write_excel_safe(
        output_path, profile["template"]["sheet_name"],
        lambda ws: apply_column_mapping(ws, df, profile["mappings"][0]),
    )

    wb = load_workbook(output_path, data_only=True)
    ws = wb[profile["template"]["sheet_name"]]
    value = effective_reader(ws, "C15")  # redirected to C10
    wb.close()

    assert value is not None
    assert abs(float(value) - 245.3) < 1e-6


def test_cell_mapping_system_date(tmp_path, profile_factory, effective_reader):
    """system_date writes today's date (dd/mm/YYYY) at the effective cell."""
    profile = profile_factory(tmp_path)
    identity = MockIdentity("ORD-C-2", "GREEN")

    output_path = prepare_output_file(profile, identity)
    write_excel_safe(
        output_path, profile["template"]["sheet_name"],
        lambda ws: apply_cell_mapping(ws, pd.DataFrame(), profile["mappings"][1], csv_path="x.csv"),
    )

    wb = load_workbook(output_path, data_only=True)
    ws = wb[profile["template"]["sheet_name"]]
    value = effective_reader(ws, "B7")  # redirected to A7
    wb.close()

    assert str(value) == datetime.now().strftime("%d/%m/%Y")


def test_cell_mapping_static(tmp_path, profile_factory, effective_reader):
    """A 'static:...' source writes the literal string at the effective cell."""
    profile = profile_factory(tmp_path)
    identity = MockIdentity("ORD-C-3", "BLUE")

    output_path = prepare_output_file(profile, identity)
    write_excel_safe(
        output_path, profile["template"]["sheet_name"],
        lambda ws: apply_cell_mapping(ws, pd.DataFrame(), profile["mappings"][2], csv_path="x.csv"),
    )

    wb = load_workbook(output_path, data_only=True)
    ws = wb[profile["template"]["sheet_name"]]
    value = effective_reader(ws, "B12")  # redirected to B10
    wb.close()

    assert value == "ISO 13934-1"


def test_identity_mapping(tmp_path, profile_factory, effective_reader):
    """Order/color identity values land at their configured (effective) cells."""
    profile = profile_factory(tmp_path)
    identity = MockIdentity("ORD-ID-9", "TEAL")

    output_path = prepare_output_file(profile, identity)
    write_excel_safe(
        output_path, profile["template"]["sheet_name"],
        lambda ws: apply_identity_mapping(ws, identity, profile["identity"]["output_cells"]),
    )

    wb = load_workbook(output_path, data_only=True)
    ws = wb[profile["template"]["sheet_name"]]
    order_val = effective_reader(ws, "B8")  # redirected to A8
    color_val = effective_reader(ws, "B9")  # plain cell
    wb.close()

    assert order_val == "ORD-ID-9"
    assert color_val == "TEAL"


def test_atomic_write_rollback(tmp_path, profile_factory):
    """If the write callback raises, the output file is left byte-for-byte unchanged."""
    profile = profile_factory(tmp_path)
    identity = MockIdentity("ORD-C-4", "YELLOW")
    output_path = prepare_output_file(profile, identity)

    original_bytes = open(output_path, "rb").read()

    def crashing_write(ws):
        ws["A1"] = "garbage that must not persist"
        raise RuntimeError("crash on purpose")

    with pytest.raises(Exception):
        write_excel_safe(output_path, profile["template"]["sheet_name"], crashing_write)

    assert open(output_path, "rb").read() == original_bytes
    # The temp working file must be cleaned up on failure.
    assert not os.path.exists(output_path + ".tmp.xlsx")


def test_never_overwrite_template(tmp_path, profile_factory, template_path):
    """Processing copies the template; the template file itself is never modified."""
    profile = profile_factory(tmp_path)
    initial_mtime = os.path.getmtime(template_path)

    df = pd.DataFrame({"Max Force (N)": [100.0]})
    identity = MockIdentity("ORD-C-5", "PURPLE")
    output_path = prepare_output_file(profile, identity)

    assert os.path.abspath(output_path) != os.path.abspath(template_path)

    write_excel_safe(
        output_path, profile["template"]["sheet_name"],
        lambda ws: apply_column_mapping(ws, df, profile["mappings"][0]),
    )

    assert os.path.getmtime(template_path) == initial_mtime


def test_prepare_output_reuses_existing_report(tmp_path, profile_factory):
    """When a matching report already exists, it is reused instead of re-copied."""
    profile = profile_factory(tmp_path)
    identity = MockIdentity("ORD-REUSE", "RED")

    first = prepare_output_file(profile, identity)
    second = prepare_output_file(profile, identity)

    assert first == second
