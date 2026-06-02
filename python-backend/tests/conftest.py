"""
Shared pytest fixtures and helpers for the TACT backend test suite.

These tests run against the *real* TACT sample files shipped in the repo's
``scratch/`` folder (real CSV + the real Excel template). If those files are
missing on a given machine, the dependent tests skip cleanly instead of failing.
"""
import os
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter

# Make the backend package importable (services.*, models.*) regardless of CWD.
BACKEND_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

# Real TACT sample data lives in <repo>/scratch.
REPO_ROOT = Path(BACKEND_ROOT).parent
SCRATCH_DIR = REPO_ROOT / "scratch"

# Concrete sample files (real exports from TACT).
ORD001_CSV = SCRATCH_DIR / "ORD001_RED_Tensile.csv"   # has Condition block + RED color
SAMPLE_CSV = SCRATCH_DIR / "sample_tact.csv"          # flat Result-style CSV
TEMPLATE_XLSX = SCRATCH_DIR / "template.xlsx"         # real QC Excel template

# Standard identity config used across the app (mirrors docs/API_CONTRACTS.md).
STANDARD_IDENTITY_CONFIG = {
    "order_source": "both",
    "color_source": "both",
    "filename_regex": r"^([A-Z0-9]+)_([A-Z]+)_",
    "filename_order_group": 1,
    "filename_color_group": 2,
    "condition_sheet": "Condition",
    "condition_keys": ["Sample name", "Submission"],
}


def read_effective_value(ws, address):
    """
    Return the value that effectively lives at ``address``.

    openpyxl stores a merged region's value only in its top-left cell; the other
    cells are read-only ``MergedCell`` objects holding ``None``. ``excel_writer``
    redirects writes to that top-left cell, so to verify a write we must read the
    top-left of whatever merged range contains ``address``.
    """
    cell = ws[address]
    if type(cell).__name__ != "MergedCell":
        return cell.value

    row, col = coordinate_to_tuple(address)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            top_left = f"{get_column_letter(rng.min_col)}{rng.min_row}"
            return ws[top_left].value
    return cell.value


@pytest.fixture(scope="session")
def scratch_dir():
    if not SCRATCH_DIR.is_dir():
        pytest.skip(f"scratch data dir not found: {SCRATCH_DIR}")
    return SCRATCH_DIR


@pytest.fixture
def ord001_csv():
    if not ORD001_CSV.exists():
        pytest.skip(f"missing sample file: {ORD001_CSV}")
    return str(ORD001_CSV)


@pytest.fixture
def sample_csv():
    if not SAMPLE_CSV.exists():
        pytest.skip(f"missing sample file: {SAMPLE_CSV}")
    return str(SAMPLE_CSV)


@pytest.fixture
def template_path():
    if not TEMPLATE_XLSX.exists():
        pytest.skip(f"missing template: {TEMPLATE_XLSX}")
    return str(TEMPLATE_XLSX)


@pytest.fixture
def template_sheet(template_path):
    """The first (and only) sheet name of the real template, fetched dynamically."""
    wb = load_workbook(template_path, read_only=True)
    try:
        return wb.sheetnames[0]
    finally:
        wb.close()


@pytest.fixture
def effective_reader():
    """Expose ``read_effective_value`` to tests as a callable fixture."""
    return read_effective_value
