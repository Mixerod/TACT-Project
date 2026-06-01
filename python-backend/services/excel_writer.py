import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Callable, Union
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import coordinate_to_tuple, get_column_letter

class ExcelWriterError(Exception):
    def __init__(self, code: str, message: str, detail: str = ""):
        super().__init__(message)
        self.code = code
        self.message = message
        self.detail = detail

def clean_value(val: Any) -> Any:
    """
    Cleans value for Excel writing:
    - NaN/NA/None -> None
    - Numeric string -> float/int
    - Regular string -> stripped string
    """
    if pd.isna(val) or val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        val_stripped = val.strip()
        try:
            # Try parsing as float
            f_val = float(val_stripped)
            # If it's an integer value represented as float, convert to int
            if f_val.is_integer():
                return int(f_val)
            return f_val
        except ValueError:
            return val_stripped
    return val

def write_to_cell_safe(ws: Worksheet, cell_address: str, value: Any) -> None:
    """
    Writes a value to a cell. If the target cell is a MergedCell (read-only),
    finds the top-left cell of the merged range containing it and writes there.
    """
    cell = ws[cell_address]
    cell_type = type(cell).__name__
    if cell_type == 'MergedCell':
        # Find the merged range containing this cell
        for merged_range in ws.merged_cells.ranges:
            row, col = coordinate_to_tuple(cell_address)
            if (merged_range.min_row <= row <= merged_range.max_row and 
                merged_range.min_col <= col <= merged_range.max_col):
                # Write to top-left cell of this range
                top_left_coord = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
                print(f"[DEBUG WRITE] {cell_address} is MergedCell. Redirected write to top-left {top_left_coord} = {value}")
                ws[top_left_coord] = value
                return
    # Not a merged cell, write normally
    print(f"[DEBUG WRITE] Writing to standard cell {cell_address} = {value}")
    ws[cell_address] = value

from services.matcher import find_existing_report, build_output_filename, format_date

def prepare_output_file(profile: Any, identity: Any) -> str:
    """
    Finds or creates output Excel file. Returns absolute path to the file.
    Supports both class objects (Pydantic) and dictionaries.
    """
    # Helper to retrieve properties from dict or object
    def get_prop(obj, key, default=None):
        if isinstance(obj, dict):
            return obj.get(key, default)
        return getattr(obj, key, default)
        
    output_cfg = get_prop(profile, 'output')
    template_cfg = get_prop(profile, 'template')
    method_code = get_prop(profile, 'method_code', '')
    
    output_dir = get_prop(output_cfg, 'directory')
    filename_pattern = get_prop(output_cfg, 'filename_pattern')
    date_format = get_prop(output_cfg, 'date_format')
    
    template_path = get_prop(template_cfg, 'path')
    
    order = get_prop(identity, 'order')
    color = get_prop(identity, 'color')
    
    # Ensure output directory exists
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    # 1. Search for existing file
    existing_path = find_existing_report(output_dir, order, color)
    if existing_path:
        return existing_path
        
    # 2. Copy template to new output file
    output_filename = build_output_filename(filename_pattern, order, color, date_format, method_code)
    output_path = Path(output_dir) / output_filename
    
    if not Path(template_path).exists():
        raise ExcelWriterError(
            code="TEMPLATE_NOT_FOUND",
            message=f"Không tìm thấy file Excel template tại đường dẫn: {template_path}",
            detail=f"Template file not found: {template_path}"
        )
        
    try:
        shutil.copy2(template_path, output_path)
    except Exception as e:
        raise ExcelWriterError(
            code="WRITE_ERROR",
            message=f"Không thể sao chép file template sang file báo cáo mới.",
            detail=str(e)
        )
        
    return str(output_path)

def write_excel_safe(output_path: str, sheet_name: str, write_fn: Callable[[Worksheet], None]) -> None:
    """
    Safely writes to Excel file using atomic write:
    Copy to tmp -> load -> write -> save -> replace.
    """
    tmp_path = output_path + ".tmp.xlsx"
    try:
        import io
        # Read the file into memory and close the file descriptor immediately
        with open(output_path, "rb") as f:
            file_bytes = f.read()
            
        # Load workbook from memory stream
        wb = load_workbook(io.BytesIO(file_bytes))
        if sheet_name not in wb.sheetnames:
            raise ExcelWriterError(
                code="SHEET_NOT_FOUND",
                message=f"Không tìm thấy sheet '{sheet_name}' trong file Excel.",
                detail=f"Sheets found: {wb.sheetnames}"
            )
            
        ws = wb[sheet_name]
        
        # Execute write operations
        write_fn(ws)
        
        # Save changes directly to the temporary path
        wb.save(tmp_path)
        wb.close()
        
        # Atomic replace with retry to bypass brief antivirus/indexer locks on Windows
        import time
        max_retries = 5
        for retry_i in range(max_retries):
            try:
                if os.path.exists(output_path):
                    os.replace(tmp_path, output_path)
                else:
                    os.rename(tmp_path, output_path)
                break
            except PermissionError as pe:
                if retry_i == max_retries - 1:
                    raise ExcelWriterError(
                        code="WRITE_ERROR",
                        message="Không thể ghi dữ liệu vì file Excel đang bị khóa bởi tiến trình khác (như chương trình quét virus hoặc Microsoft Excel). Vui lòng thử lại.",
                        detail=str(pe)
                    )
                time.sleep(0.2)
            
    except ExcelWriterError as e:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except:
                pass
        raise e
    except Exception as e:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except:
                pass
        raise ExcelWriterError(
            code="WRITE_ERROR",
            message="Đã xảy ra lỗi khi ghi dữ liệu báo cáo Excel.",
            detail=str(e)
        )

def apply_column_mapping(ws: Worksheet, df: pd.DataFrame, mapping: Dict[str, Any]) -> None:
    """
    Applies column mapping from CSV/Excel DataFrame to Excel Worksheet.
    """
    col_letter = mapping.get('excel_column')
    start_row = mapping.get('excel_start_row')
    csv_col = mapping.get('csv_column')
    
    if not col_letter or start_row is None or not csv_col:
        raise ExcelWriterError(
            code="INVALID_MAPPING",
            message="Cấu hình Column Mapping không hợp lệ.",
            detail=f"Mapping config: {mapping}"
        )
        
    if csv_col not in df.columns:
        raise ExcelWriterError(
            code="INVALID_MAPPING",
            message=f"Cột dữ liệu '{csv_col}' không tồn tại trong file dữ liệu gốc.",
            detail=f"CSV Columns: {list(df.columns)}"
        )
        
    values = df[csv_col].tolist()
    for i, val in enumerate(values):
        cell_addr = f"{col_letter}{start_row + i}"
        write_to_cell_safe(ws, cell_addr, clean_value(val))

def apply_cell_mapping(ws: Worksheet, df: pd.DataFrame, mapping: Dict[str, Any], csv_path: str) -> None:
    """
    Applies cell mapping from various sources to Excel Worksheet.
    """
    excel_cell = mapping.get('excel_cell')
    value_source = mapping.get('value_source')
    
    if not excel_cell or not value_source:
        raise ExcelWriterError(
            code="INVALID_MAPPING",
            message="Cấu hình Cell Mapping không hợp lệ.",
            detail=f"Mapping config: {mapping}"
        )
        
    # Resolve value_source
    val = None
    if value_source == "system_date":
        val = datetime.now().strftime("%d/%m/%Y")
    elif value_source == "csv_filename":
        val = Path(csv_path).stem
    elif str(value_source).startswith("static:"):
        val = str(value_source)[7:]
    else:
        # Otherwise treat it as a column name and get first row
        if value_source in df.columns:
            if len(df) > 0:
                val = df[value_source].iloc[0]
            else:
                val = None
        else:
            raise ExcelWriterError(
                code="INVALID_MAPPING",
                message=f"Nguồn giá trị ô '{value_source}' không khớp với các biến hệ thống hay cột dữ liệu gốc.",
                detail=f"value_source: {value_source}, CSV Columns: {list(df.columns)}"
            )
            
    write_to_cell_safe(ws, excel_cell, clean_value(val))

def apply_range_mapping(ws: Worksheet, df: pd.DataFrame, mapping: Dict[str, Any]) -> None:
    """
    Applies range mapping by copying block data from DataFrame to Excel Worksheet.
    """
    start_cell = mapping.get('excel_start_cell')
    csv_columns = mapping.get('csv_columns')
    
    if not start_cell or not csv_columns or not isinstance(csv_columns, list):
        raise ExcelWriterError(
            code="INVALID_MAPPING",
            message="Cấu hình Range Mapping không hợp lệ.",
            detail=f"Mapping config: {mapping}"
        )
        
    for col in csv_columns:
        if col not in df.columns:
            raise ExcelWriterError(
                code="INVALID_MAPPING",
                message=f"Cột dữ liệu '{col}' trong Range Mapping không tồn tại trong file gốc.",
                detail=f"Missing column: {col}, CSV Columns: {list(df.columns)}"
            )
            
    try:
        # Parse start cell using openpyxl coordinate_to_tuple -> returns (row, col)
        start_row, start_col_idx = coordinate_to_tuple(start_cell)
    except Exception as e:
        raise ExcelWriterError(
            code="INVALID_MAPPING",
            message=f"Địa chỉ ô bắt đầu '{start_cell}' không hợp lệ.",
            detail=str(e)
        )
        
    # Write block
    for row_i, (_, row) in enumerate(df[csv_columns].iterrows()):
        for col_i, csv_col in enumerate(csv_columns):
            current_col_letter = get_column_letter(start_col_idx + col_i)
            cell_addr = f"{current_col_letter}{start_row + row_i}"
            write_to_cell_safe(ws, cell_addr, clean_value(row[csv_col]))

def apply_identity_mapping(ws: Worksheet, identity_result: Any, output_cells: List[Dict[str, Any]]) -> None:
    """
    Applies order ID and color identity values to specified cells in Excel Worksheet.
    """
    # Helper to retrieve properties from dict or object
    def get_prop(obj, key, default=None):
        if isinstance(obj, dict):
            return obj.get(key, default)
        return getattr(obj, key, default)
        
    order = get_prop(identity_result, 'order')
    color = get_prop(identity_result, 'color')
    
    field_values = {
        "order": order,
        "color": color
    }
    
    for cell_config in output_cells:
        field = get_prop(cell_config, 'field')
        cell = get_prop(cell_config, 'cell')
        if not field or not cell:
            continue
            
        value = field_values.get(field)
        if value is not None:
            write_to_cell_safe(ws, cell, clean_value(value))

def read_excel_preview(file_path: str, sheet_name: str = None) -> Dict[str, Any]:
    """
    Reads an Excel file and returns structural grid preview data up to 30 rows and 10 columns.
    """
    path = Path(file_path)
    if not path.exists():
        raise ExcelWriterError(
            code="FILE_NOT_FOUND",
            message=f"Không tìm thấy file Excel template tại đường dẫn: '{file_path}'",
            detail=f"Excel file not found: {file_path}"
        )
        
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        raise ExcelWriterError(
            code="PARSE_ERROR",
            message="Không thể đọc hoặc phân tích file Excel template.",
            detail=str(e)
        )
        
    sheets = wb.sheetnames
    if not sheet_name or sheet_name not in sheets:
        sheet_name = sheets[0] if sheets else "Sheet1"
        
    ws = wb[sheet_name]
    
    # Grid limits: max 30 rows, 10 columns
    max_r = min(ws.max_row, 30) if ws.max_row else 30
    max_c = min(ws.max_column, 10) if ws.max_column else 10
    
    # Ensure a minimum size
    max_r = max(max_r, 10)
    max_c = max(max_c, 5)
    
    cells_grid = []
    for r in range(1, max_r + 1):
        row_cells = []
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            
            is_empty = val is None or str(val).strip() == ""
            
            style_hint = 'empty'
            if is_empty:
                style_hint = 'empty'
            elif cell.font and cell.font.bold:
                style_hint = 'header'
            elif cell.data_type == 'f' or (isinstance(val, str) and str(val).startswith('=')):
                style_hint = 'formula'
            else:
                style_hint = 'data'
                
            row_cells.append({
                "address": cell.coordinate,
                "value": val,
                "row": r,
                "col": c,
                "col_letter": get_column_letter(c),
                "is_empty": is_empty,
                "style_hint": style_hint
            })
        cells_grid.append(row_cells)
        
    wb.close()
    
    return {
        "sheet_name": sheet_name,
        "sheets": sheets,
        "cells": cells_grid,
        "row_count": max_r,
        "col_count": max_c
    }
