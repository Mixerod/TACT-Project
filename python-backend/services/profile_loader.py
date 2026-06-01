import json
from pathlib import Path
from typing import Dict, List, Any, Tuple
from pydantic import ValidationError
from models.schemas import Profile
from openpyxl import load_workbook

class ProfileValidationError(Exception):
    def __init__(self, errors: List[Dict[str, str]], warnings: List[Dict[str, str]]):
        self.errors = errors
        self.warnings = warnings
        super().__init__("Profile validation failed")

def load_profile(file_path: str) -> Profile:
    """
    Loads and validates a Profile JSON file.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Profile file not found at: {file_path}")
        
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
            
        return Profile.model_validate(data)
    except ValidationError as ve:
        # Reformat Pydantic validation errors
        errors = []
        for err in ve.errors():
            loc = ".".join(str(x) for x in err["loc"])
            errors.append({
                "field": loc,
                "message": f"Dữ liệu không hợp lệ: {err['msg']}"
            })
        raise ProfileValidationError(errors=errors, warnings=[])
    except Exception as e:
        raise ProfileValidationError(
            errors=[{"field": "json", "message": f"Không thể đọc file JSON profile: {str(e)}"}],
            warnings=[]
        )

def save_profile(file_path: str, profile: Profile) -> None:
    """
    Saves a Profile Pydantic model to a JSON file.
    """
    path = Path(file_path)
    # Ensure parent folder exists
    path.parent.mkdir(parents=True, exist_ok=True)
    
    # Convert Profile to dict, serializing correctly
    data = profile.model_dump()
    
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def validate_profile_config(profile_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Validates a profile configuration.
    Performs structural Pydantic validation followed by file system sanity checks.
    """
    errors = []
    warnings = []
    profile = None
    
    # 1. Structural Schema Validation
    try:
        profile = Profile.model_validate(profile_data)
    except ValidationError as ve:
        for err in ve.errors():
            loc = ".".join(str(x) for x in err["loc"])
            errors.append({
                "field": loc,
                "message": f"Cấu hình trường '{loc}' không hợp lệ: {err['msg']}"
            })
            
    if len(errors) > 0:
        return {
            "valid": False,
            "errors": errors,
            "warnings": warnings
        }
        
    # 2. File System and Template validation
    # Template Path exists
    template_path_str = profile.template.path
    template_path = Path(template_path_str)
    if not template_path.exists():
        errors.append({
            "field": "template.path",
            "message": f"File template Excel không tồn tại tại đường dẫn: '{template_path_str}'"
        })
    else:
        # Template Sheet exists
        try:
            wb = load_workbook(template_path_str, read_only=True)
            if profile.template.sheet_name not in wb.sheetnames:
                errors.append({
                    "field": "template.sheet_name",
                    "message": f"Sheet '{profile.template.sheet_name}' không tồn tại trong file template Excel. Các sheet hiện có: {wb.sheetnames}"
                })
            wb.close()
        except Exception as e:
            errors.append({
                "field": "template.path",
                "message": f"Không thể mở file template Excel: {str(e)}"
            })
            
    # Output Directory exists (warn if not)
    output_dir_str = profile.output.directory
    output_dir = Path(output_dir_str)
    if not output_dir.exists():
        warnings.append({
            "field": "output.directory",
            "message": f"Thư mục lưu báo cáo '{output_dir_str}' chưa tồn tại. Thư mục này sẽ tự động được tạo khi process chạy."
        })
        
    return {
        "valid": len(errors) == 0,
        "errors": errors,
        "warnings": warnings
    }

def load_profile_by_id(profile_id: str) -> Profile:
    """
    Finds and loads a Profile by its ID from APPDATA or local profiles folders.
    """
    import os
    app_data = os.environ.get("APPDATA")
    if not app_data:
        app_data = str(Path.home() / "AppData" / "Roaming")
        
    profiles_dir = Path(app_data) / "TACTAutomation" / "profiles"
    local_profiles_dir = Path(__file__).parent.parent.parent / "profiles"
    
    # Try direct mapping
    path = profiles_dir / f"{profile_id}.json"
    if path.exists():
        return load_profile(str(path))
        
    path_local = local_profiles_dir / f"{profile_id}.json"
    if path_local.exists():
        return load_profile(str(path_local))
        
    # Search all JSON files in roaming profiles dir
    if profiles_dir.exists():
        for f in profiles_dir.glob("*.json"):
            try:
                prof = load_profile(str(f))
                if prof.id == profile_id:
                    return prof
            except:
                pass
                
    # Search all JSON files in local profiles dir
    if local_profiles_dir.exists():
        for f in local_profiles_dir.glob("*.json"):
            try:
                prof = load_profile(str(f))
                if prof.id == profile_id:
                    return prof
            except:
                pass
                
    raise FileNotFoundError(f"Không tìm thấy cấu hình Profile có ID: '{profile_id}'")
